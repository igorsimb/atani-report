from openpyxl import load_workbook


# pylint: disable=too-many-instance-attributes
class File:
    def __init__(self, file_path):
        self.report_wb = load_workbook(file_path, read_only=True, data_only=True)

        # svod_shop
        self.svod_shop = self.report_wb.worksheets[3]

        # Дата
        self.date_range = self.svod_shop["E1"].value

        # Заказано товаров, шт.
        self.ordered_items_text = self.svod_shop["A5"].value
        self.ordered_items_value = self.svod_shop["E5"].value
        self.ordered_items_value = f"{self.ordered_items_value:,}".replace(",", " ")

        # Заказано на сумму, руб.
        self.ordered_sum_text = self.svod_shop["A4"].value
        self.ordered_sum_value = self.svod_shop["E4"].value
        self.ordered_sum_value = f"{self.ordered_sum_value:,}".replace(",", " ")

        # Продажи, шт.
        self.sold_items_text = self.svod_shop["A10"].value
        self.sold_items_value = self.svod_shop["E10"].value
        self.sold_items_value = f"{self.sold_items_value:,}".replace(",", " ")

        # Продажи, руб.
        self.sold_sum_text = self.svod_shop["A9"].value
        self.sold_sum_value = self.svod_shop["E9"].value
        self.sold_sum_value = f"{self.sold_sum_value:,}".replace(",", " ")

        # Средняя цена продажи, руб.
        self.ave_sum_text = self.svod_shop["A11"].value
        self.ave_sum_value = self.svod_shop["E11"].value
        self.ave_sum_value = f"{self.ave_sum_value:,.2f}".replace(",", " ")

        # CTR, %
        self.ctr_text = self.svod_shop["A29"].value
        self.ctr_value = f'{self.svod_shop["E29"].value * 100:.2f}'

        # ДРР, %
        self.drr_text = self.svod_shop["A35"].value
        self.drr_value = f'{self.svod_shop["E35"].value * 100:.2f}'


def get_column_index_by_cell_value(sheet, cell_value):
    result_column = 0
    for column in sheet.iter_rows(min_row=1, max_row=1, max_col=10):
        for cell in column:
            if str(cell.value).lower() == cell_value.lower():
                print(f"Your column is #{cell.column} (i.e. column {cell.column_letter})")
                result_column = cell.column
                break
        else:
            continue
        break
    return result_column


def overview(file):
    marketplace = "ОЗОН"
    overview = f"Высылаем отчет за {file.date_range}\n" \
               f"{marketplace}\n\n" \
               f"🛒 Заказано в шт {file.ordered_items_value} на сумму {file.ordered_sum_value} руб\n" \
               f"✅ Продажи в шт {file.sold_items_value} на сумму {file.sold_sum_value}р.\n" \
               f"Средняя цена продажи {file.ave_sum_value} р\n" \
               f"CTR {file.ctr_value}%\n" \
               f"ДРР {file.drr_value}%\n"
    return overview


def conclusions(file):
    # ВЫВОДЫ
    conclusions_sheet = file.report_wb.worksheets[0]

    conclusions_column = get_column_index_by_cell_value(sheet=conclusions_sheet, cell_value="Выводы")
    conc_text = conclusions_sheet.iter_rows(min_row=2, min_col=conclusions_column, max_col=conclusions_column)

    conc_text_list = []

    # create a list of values from conc_text generator
    for row in conc_text:
        conc_text_list.extend(f"❕{cell.value}" for cell in row if cell.value is not None)

    # turn list into string with separate lines
    return "\n".join(conc_text_list)


def plan_for_next_week(file):
    # ПЛАН
    plan_sheet = file.report_wb.worksheets[1]
    plan_text = plan_sheet.iter_rows(min_row=2, min_col=2, max_col=2)  # column B
    plan_status = plan_sheet.iter_rows(min_row=2, min_col=3, max_col=3)  # column C

    # create lists from plan_text and plan_status generators
    plan_text_list = []
    plan_status_list = []

    for index, row in enumerate(plan_text, start=1):
        plan_text_list.extend(cell.value for cell in row if cell.value is not None)
    for index, row in enumerate(plan_status, start=1):
        plan_status_list.extend(cell.value for cell in row if cell.value is not None)

    # join the lists into one and enumerate the items
    plan_list = ["\n✅ Задачи на неделю:\n"]
    for index, value in enumerate(zip(plan_text_list, plan_status_list), start=1):
        plan_list.append(f"{index}. {value[0]}:    {value[1]}")

    # turn list into string with separate lines
    return "\n".join(plan_list)


def raw_data(file):
    file = File(file)
    log_list = [
        f"Sheet name: {file.svod_shop.title}\n",
        f"[A5]: {file.ordered_items_text}\n[E5]: {file.ordered_items_value}\n",
        f"[A4]: {file.ordered_sum_text}\n[E4]: {file.ordered_sum_value}\n",
        f"[A10]: {file.sold_items_text}\n[E10]: {file.sold_items_value}\n",
        f"[A9]: {file.sold_sum_text}\n[E9]: {file.sold_sum_value}\n",
        f"[A11]: {file.ave_sum_text}\n[E11]: {file.ave_sum_value}\n",
        f"[A29]: {file.ctr_text}\n[E29]: {file.ctr_value}\n",
        f"[A35]: {file.drr_text}\n[E35]: {file.drr_value}",
    ]
    return "\n".join(log_list)


def final_report(file):
    # instantiating class File so that generate_report function only needs the file path value
    file = File(file)
    return f"{overview(file)}\n{conclusions(file)}\n{plan_for_next_week(file)}"
