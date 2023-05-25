from openpyxl import load_workbook


# pylint: disable=too-many-instance-attributes
class OzonFile:
    def __init__(self, file_path):
        self.report_wb = load_workbook(file_path, read_only=True, data_only=True)

        # svod_shop
        self.svod_shop = self.report_wb.worksheets[3]

        # Дата
        self.date_range = self.svod_shop["E1"].value

        # Заказано на сумму, руб.
        self.ordered_sum_text = self.svod_shop["A4"].value
        self.ordered_sum_value = self.svod_shop["E4"].value
        self.ordered_sum_value = f"{self.ordered_sum_value:,}".replace(",", " ")

        # Заказано товаров, шт.
        self.ordered_items_text = self.svod_shop["A5"].value
        self.ordered_items_value = self.svod_shop["E5"].value
        self.ordered_items_value = f"{self.ordered_items_value:,}".replace(",", " ")

        # Продажи, руб.
        self.sold_sum_text = self.svod_shop["A9"].value
        self.sold_sum_value = self.svod_shop["E9"].value
        self.sold_sum_value = f"{self.sold_sum_value:,}".replace(",", " ")

        # Продажи, шт.
        self.sold_items_text = self.svod_shop["A10"].value
        self.sold_items_value = self.svod_shop["E10"].value
        self.sold_items_value = f"{self.sold_items_value:,}".replace(",", " ")

        # Средняя цена продажи, руб.
        self.ave_sum_text = self.svod_shop["A11"].value
        self.ave_sum_value = self.svod_shop["E11"].value
        self.ave_sum_value = f"{self.ave_sum_value:,.2f}".replace(",", " ")

        # CTR, %
        self.ctr_text = self.svod_shop["A29"].value
        self.ctr_value = f'{self.svod_shop["E29"].value * 100:.2f}'

        # ДРР, %
        self.drr_text = self.svod_shop["A35"].value
        self.drr_value = f'{self.svod_shop["E35"].value * 100:.2f}'


class WbFile:
    def __init__(self, file_path):
        self.report_wb = load_workbook(file_path, read_only=True, data_only=True)

        # svod_shop
        self.svod_shop = self.report_wb.worksheets[3]

        # Дата
        self.date_range = self.svod_shop["E1"].value

        # Заказано на сумму, руб.
        self.ordered_sum_text = self.svod_shop["A2"].value
        self.ordered_sum_value = self.svod_shop["E2"].value
        self.ordered_sum_value = f"{self.ordered_sum_value:,}".replace(",", " ")

        # Заказано товаров, шт.
        self.ordered_items_text = self.svod_shop["A3"].value
        self.ordered_items_value = self.svod_shop["E3"].value
        self.ordered_items_value = f"{self.ordered_items_value:,}".replace(",", " ")

        # Продажи, руб.
        self.sold_sum_text = self.svod_shop["A5"].value
        self.sold_sum_value = self.svod_shop["E5"].value
        self.sold_sum_value = f"{self.sold_sum_value:,}".replace(",", " ")

        # Продажи, шт.
        self.sold_items_text = self.svod_shop["A6"].value
        self.sold_items_value = self.svod_shop["E6"].value
        self.sold_items_value = f"{self.sold_items_value:,}".replace(",", " ")

        # Средняя цена продажи, руб.
        self.ave_sum_text = self.svod_shop["A9"].value
        self.ave_sum_value = self.svod_shop["E9"].value
        self.ave_sum_value = f"{self.ave_sum_value:,.2f}".replace(",", " ")

        # CTR, %
        self.ctr_text = self.svod_shop["A16"].value
        self.ctr_value = f'{self.svod_shop["E16"].value * 100:.2f}'

        # ДРР, %
        self.drr_text = self.svod_shop["A18"].value
        self.drr_value = f'{self.svod_shop["E18"].value * 100:.2f}'


class YandexFile:
    def __init__(self, file_path):
        self.report_wb = load_workbook(file_path, read_only=True, data_only=True)

        # "ВЫВОДЫ" sheet may be absent is some reports. This checks if it exists and assigns svod_shop accordingly.
        if "ВЫВОДЫ" in self.report_wb.sheetnames:
            self.svod_shop = self.report_wb.worksheets[1]
        else:
            self.svod_shop = self.report_wb.worksheets[0]

        # Дата
        self.date_range = self.svod_shop["E1"].value

        # Заказано на сумму, руб.
        self.ordered_sum_text = self.svod_shop["A3"].value
        self.ordered_sum_value = self.svod_shop["E3"].value
        self.ordered_sum_value = f"{self.ordered_sum_value:,}".replace(",", " ")

        # Заказано товаров, шт.
        self.ordered_items_text = self.svod_shop["A4"].value
        self.ordered_items_value = self.svod_shop["E4"].value
        self.ordered_items_value = f"{self.ordered_items_value:,}".replace(",", " ")

        # Продажи, руб.
        self.sold_sum_text = self.svod_shop["A9"].value
        self.sold_sum_value = self.svod_shop["E9"].value
        self.sold_sum_value = f"{self.sold_sum_value:,}".replace(",", " ")

        # Продажи, шт.
        self.sold_items_text = self.svod_shop["A10"].value
        self.sold_items_value = self.svod_shop["E10"].value
        self.sold_items_value = f"{self.sold_items_value:,}".replace(",", " ")

        # Средняя цена продажи, руб.
        self.ave_sum_text = "Средняя цена продажи, руб."
        self.ave_sum_value = int(float(self.sold_sum_value.replace(' ', ''))/float(self.sold_items_value))
        self.ave_sum_value = f"{self.ave_sum_value}".replace(",", " ")

        # CTR, %
        self.ctr_text = None
        self.ctr_value = None

        # ДРР, %
        self.drr_text = self.svod_shop["A22"].value # A22
        self.drr_value = f'{self.svod_shop["E22"].value * 100:.2f}'
